#!/usr/bin/env python3
"""
Export Vector Database Chunks to Excel
Creates an Excel file with all chunks for easy viewing and analysis
"""
import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.utils.vector_store import get_vector_store


def export_chunks_to_excel(file_id: str = None, output_file: str = None):
    """
    Export chunks to Excel file

    Args:
        file_id: Optional file ID to filter by specific document
        output_file: Output Excel filename (default: chunks_export_TIMESTAMP.xlsx)
    """
    vector_store = get_vector_store()

    # Get all data from the collection
    all_data = vector_store.collection.get(
        where={"file_id": file_id} if file_id else None,
        include=["documents", "metadatas", "embeddings"]
    )

    if not all_data['ids']:
        print(f"❌ No chunks found{f' for file_id: {file_id}' if file_id else ''}")
        return

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group chunks by file_id
    files = {}
    for i, chunk_id in enumerate(all_data['ids']):
        metadata = all_data['metadatas'][i]
        doc_file_id = metadata.get('file_id', 'unknown')

        if doc_file_id not in files:
            files[doc_file_id] = {
                'filename': metadata.get('filename', 'unknown'),
                'format': metadata.get('format', 'unknown'),
                'chunks': []
            }

        files[doc_file_id]['chunks'].append({
            'chunk_id': chunk_id,
            'chunk_index': metadata.get('chunk_index', 0),
            'total_chunks': metadata.get('total_chunks', 0),
            'text': all_data['documents'][i],
            'text_length': len(all_data['documents'][i]),
            'embedding_dims': len(all_data['embeddings'][i]) if all_data['embeddings'] else 0,
            'metadata': metadata
        })

    # Create sheets for each document
    for doc_file_id, file_data in files.items():
        # Sort chunks by index
        chunks = sorted(file_data['chunks'], key=lambda x: x['chunk_index'])

        # Create sheet (Excel sheet names max 31 chars)
        sheet_name = file_data['filename'][:28] + "..." if len(file_data['filename']) > 31 else file_data['filename']
        ws = wb.create_sheet(title=sheet_name)

        # Header row styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Set up headers
        headers = [
            ("Chunk #", 10),
            ("Chunk ID", 35),
            ("Text Length", 12),
            ("Embedding Dims", 15),
            ("Chunk Text", 100)
        ]

        for col_num, (header, width) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            ws.column_dimensions[get_column_letter(col_num)].width = width

        # Add document info above headers
        ws.insert_rows(1, 3)
        info_font = Font(bold=True, size=11)
        ws.cell(row=1, column=1, value="Document:").font = info_font
        ws.cell(row=1, column=2, value=file_data['filename'])
        ws.cell(row=2, column=1, value="File ID:").font = info_font
        ws.cell(row=2, column=2, value=doc_file_id)
        ws.cell(row=3, column=1, value="Total Chunks:").font = info_font
        ws.cell(row=3, column=2, value=len(chunks))

        # Add data rows (starting from row 5, after headers)
        for idx, chunk in enumerate(chunks, start=5):
            # Chunk number
            ws.cell(row=idx, column=1, value=chunk['chunk_index'] + 1).alignment = Alignment(horizontal='center')

            # Chunk ID
            ws.cell(row=idx, column=2, value=chunk['chunk_id'])

            # Text length
            ws.cell(row=idx, column=3, value=chunk['text_length']).alignment = Alignment(horizontal='center')

            # Embedding dimensions
            ws.cell(row=idx, column=4, value=chunk['embedding_dims']).alignment = Alignment(horizontal='center')

            # Chunk text (wrap text for readability)
            text_cell = ws.cell(row=idx, column=5, value=chunk['text'])
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Add border to all cells
            for col in range(1, 6):
                ws.cell(row=idx, column=col).border = border

        # Set row height for better readability
        for row in range(5, 5 + len(chunks)):
            ws.row_dimensions[row].height = 100  # Adjust height for wrapped text

        # Freeze panes (freeze header row)
        ws.freeze_panes = 'A5'

    # Create summary sheet
    summary = wb.create_sheet(title="Summary", index=0)
    summary_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    # Summary headers
    summary.cell(row=1, column=1, value="VECTOR DATABASE SUMMARY").font = Font(bold=True, size=14)
    summary.merge_cells('A1:D1')

    summary.cell(row=3, column=1, value="Document").font = Font(bold=True)
    summary.cell(row=3, column=2, value="File ID").font = Font(bold=True)
    summary.cell(row=3, column=3, value="Format").font = Font(bold=True)
    summary.cell(row=3, column=4, value="Chunks").font = Font(bold=True)

    for col in range(1, 5):
        summary.cell(row=3, column=col).fill = summary_fill
        summary.cell(row=3, column=col).font = Font(bold=True, color="FFFFFF")

    # Add document summaries
    row = 4
    total_chunks = 0
    for doc_file_id, file_data in files.items():
        summary.cell(row=row, column=1, value=file_data['filename'])
        summary.cell(row=row, column=2, value=doc_file_id)
        summary.cell(row=row, column=3, value=file_data['format'])
        summary.cell(row=row, column=4, value=len(file_data['chunks']))
        total_chunks += len(file_data['chunks'])
        row += 1

    # Totals
    summary.cell(row=row + 1, column=1, value="TOTAL:").font = Font(bold=True)
    summary.cell(row=row + 1, column=4, value=total_chunks).font = Font(bold=True)

    # Column widths
    summary.column_dimensions['A'].width = 40
    summary.column_dimensions['B'].width = 40
    summary.column_dimensions['C'].width = 15
    summary.column_dimensions['D'].width = 15

    # Generate output filename
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        if file_id:
            output_file = f"chunks_{file_id[:8]}_{timestamp}.xlsx"
        else:
            output_file = f"chunks_export_{timestamp}.xlsx"

    # Save workbook
    output_path = Path(output_file)
    wb.save(output_path)

    print(f"\n✅ Chunks exported successfully!")
    print(f"📄 File: {output_path.absolute()}")
    print(f"📊 Documents: {len(files)}")
    print(f"📦 Total chunks: {total_chunks}")
    print(f"\nOpen the file to view all chunks in Excel format.")


def list_documents():
    """List all documents in the vector database"""
    vector_store = get_vector_store()

    all_data = vector_store.collection.get(include=["metadatas"])

    if not all_data['ids']:
        print("❌ No documents found in vector database")
        return

    # Group by file_id
    files = {}
    for metadata in all_data['metadatas']:
        file_id = metadata.get('file_id', 'unknown')
        if file_id not in files:
            files[file_id] = {
                'filename': metadata.get('filename', 'unknown'),
                'format': metadata.get('format', 'unknown'),
                'chunk_count': 0
            }
        files[file_id]['chunk_count'] += 1

    print(f"\n{'='*80}")
    print(f"📚 DOCUMENTS IN VECTOR DATABASE")
    print(f"{'='*80}\n")

    for file_id, data in files.items():
        print(f"📄 {data['filename']}")
        print(f"   File ID: {file_id}")
        print(f"   Format: {data['format']}")
        print(f"   Chunks: {data['chunk_count']}")
        print(f"{'─'*80}")

    print(f"\nTotal: {len(files)} documents, {len(all_data['ids'])} chunks")
    print(f"\nUsage:")
    print(f"  Export all:            python export_chunks_to_excel.py")
    print(f"  Export specific doc:   python export_chunks_to_excel.py <file_id>")
    print(f"  Custom output name:    python export_chunks_to_excel.py <file_id> output.xlsx")
    print(f"{'='*80}\n")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        if sys.argv[1] in ["list", "-l", "--list"]:
            # List all documents
            list_documents()
        else:
            # Export specific document or all
            file_id = sys.argv[1] if sys.argv[1] != "all" else None
            output_file = sys.argv[2] if len(sys.argv) > 2 else None
            export_chunks_to_excel(file_id, output_file)
    else:
        # Export all documents by default
        export_chunks_to_excel()