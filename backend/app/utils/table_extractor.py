import pdfplumber
from docx import Document
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from typing import List, Dict
import re


class TableExtractor:
    """Extract tables from documents and export to Excel"""

    @staticmethod
    def extract_tables_from_pdf(file_path: str) -> List[pd.DataFrame]:
        """Extract tables from PDF using pdfplumber"""
        tables = []

        try:
            with pdfplumber.open(file_path) as pdf:
                for page_num, page in enumerate(pdf.pages, 1):
                    page_tables = page.extract_tables()

                    for table_num, table in enumerate(page_tables):
                        if table and len(table) > 0:
                            # Convert to DataFrame
                            df = pd.DataFrame(table[1:], columns=table[0] if table[0] else None)
                            df.attrs['page'] = page_num
                            df.attrs['table_num'] = table_num + 1
                            tables.append(df)

        except Exception as e:
            print(f"Error extracting tables from PDF: {str(e)}")

        return tables

    @staticmethod
    def extract_tables_from_docx(file_path: str) -> List[pd.DataFrame]:
        """Extract tables from DOCX"""
        tables = []

        try:
            doc = Document(file_path)

            for table_num, table in enumerate(doc.tables, 1):
                data = []
                for row in table.rows:
                    row_data = [cell.text.strip() for cell in row.cells]
                    data.append(row_data)

                if data and len(data) > 1:
                    # Use first row as header if it looks like a header
                    df = pd.DataFrame(data[1:], columns=data[0])
                    df.attrs['table_num'] = table_num
                    tables.append(df)

        except Exception as e:
            print(f"Error extracting tables from DOCX: {str(e)}")

        return tables

    @staticmethod
    def extract_tables_from_txt(file_path: str) -> List[pd.DataFrame]:
        """Extract table-like structures from TXT files"""
        tables = []

        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()

            # Try to find table patterns (lines with multiple delimiters)
            lines = content.split('\n')
            current_table = []
            table_num = 0

            for line in lines:
                # Detect table rows (lines with tabs, pipes, or multiple spaces)
                if '\t' in line or '|' in line or re.search(r'\s{2,}', line):
                    current_table.append(line)
                else:
                    if len(current_table) > 2:  # At least header + 2 rows
                        table_num += 1
                        # Try to parse the table
                        df = TableExtractor._parse_text_table(current_table)
                        if df is not None:
                            df.attrs['table_num'] = table_num
                            tables.append(df)
                    current_table = []

            # Check last table
            if len(current_table) > 2:
                table_num += 1
                df = TableExtractor._parse_text_table(current_table)
                if df is not None:
                    df.attrs['table_num'] = table_num
                    tables.append(df)

        except Exception as e:
            print(f"Error extracting tables from TXT: {str(e)}")

        return tables

    @staticmethod
    def _parse_text_table(lines: List[str]) -> pd.DataFrame:
        """Parse text lines into DataFrame"""
        try:
            data = []
            for line in lines:
                # Split by tabs, pipes, or multiple spaces
                if '\t' in line:
                    row = [cell.strip() for cell in line.split('\t')]
                elif '|' in line:
                    row = [cell.strip() for cell in line.split('|') if cell.strip()]
                else:
                    row = [cell.strip() for cell in re.split(r'\s{2,}', line) if cell.strip()]

                if row:
                    data.append(row)

            if len(data) > 1:
                return pd.DataFrame(data[1:], columns=data[0])
            return None
        except:
            return None

    @staticmethod
    def export_to_excel(tables: List[pd.DataFrame], output_path: str, filename: str = "extracted_tables.xlsx") -> Dict:
        """Export extracted tables to Excel with formatting"""
        try:
            full_path = Path(output_path) / filename
            full_path.parent.mkdir(parents=True, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            for idx, df in enumerate(tables, 1):
                # Create sheet name
                page = df.attrs.get('page', '')
                table_num = df.attrs.get('table_num', idx)
                sheet_name = f"Table_{idx}" if not page else f"P{page}_T{table_num}"
                sheet_name = sheet_name[:31]  # Excel limit

                ws = wb.create_sheet(sheet_name)

                # Write headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col_idx, column in enumerate(df.columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=str(column))
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Write data
                for row_idx, row in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=str(value))
                        cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(full_path)

            return {
                "success": True,
                "file_path": str(full_path),
                "tables_count": len(tables),
                "filename": filename
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"Excel export failed: {str(e)}"
            }

    @staticmethod
    def extract_and_export(file_path: str, output_dir: str) -> Dict:
        """Main method to extract tables and export to Excel"""
        file_extension = Path(file_path).suffix.lower()
        tables = []

        if file_extension == '.pdf':
            tables = TableExtractor.extract_tables_from_pdf(file_path)
        elif file_extension in ['.docx', '.doc']:
            tables = TableExtractor.extract_tables_from_docx(file_path)
        elif file_extension == '.txt':
            tables = TableExtractor.extract_tables_from_txt(file_path)

        if not tables:
            return {
                "success": False,
                "error": "No tables found in document",
                "tables_count": 0
            }

        # Generate filename
        original_name = Path(file_path).stem
        excel_filename = f"{original_name}_tables.xlsx"

        # Export to Excel
        result = TableExtractor.export_to_excel(tables, output_dir, excel_filename)

        return result
